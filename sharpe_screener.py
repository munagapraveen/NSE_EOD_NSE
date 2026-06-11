"""Sharpe ratio screener for the standalone NSE EOD database."""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

from db import get_connection, setup_schema
from logger import get_logger

log = get_logger(__name__)

MCAP_FILTER_CR = 1000.0
ROC_ANNUAL_FILTER = 6.5
TURNOVER_FILTER_CR = 1.0
TOP_N = 50
ROC_3M_FILTER = 20.0
CIRCUIT_MAX_HITS = 10
CIRCUIT_LOOKBACK = 63
CIRCUIT_BANDS = (5.0, 10.0, 20.0)
CIRCUIT_TOLERANCE = 0.025
TRADING_DAYS_52W = 252
DAYS_TO_LOAD = 380


def round2(value):
    if value is None or pd.isna(value):
        return None
    return round(float(value), 2)


def parse_args(args):
    options = {
        "mcap": MCAP_FILTER_CR,
        "rf": ROC_ANNUAL_FILTER,
        "turnover": TURNOVER_FILTER_CR,
        "top": TOP_N,
        "date": None,
        "long_months": 6,
        "short_months": 3,
    }
    for i, arg in enumerate(args):
        if arg == "--mcap" and i + 1 < len(args):
            options["mcap"] = float(args[i + 1])
        if arg == "--rf" and i + 1 < len(args):
            options["rf"] = float(args[i + 1])
        if arg == "--turnover" and i + 1 < len(args):
            options["turnover"] = float(args[i + 1])
        if arg == "--top" and i + 1 < len(args):
            options["top"] = int(args[i + 1])
        if arg == "--date" and i + 1 < len(args):
            options["date"] = args[i + 1].strip()
        if arg == "--long-months" and i + 1 < len(args):
            options["long_months"] = int(args[i + 1])
        if arg == "--short-months" and i + 1 < len(args):
            options["short_months"] = int(args[i + 1])
    return options


def resolve_snapshot_date(conn, as_of_date=None):
    if as_of_date:
        snapshot_date = conn.execute(
            "SELECT MAX(date) FROM adjusted_eod_prices WHERE date <= ?",
            (as_of_date,),
        ).fetchone()[0]
    else:
        snapshot_date = conn.execute("SELECT MAX(date) FROM adjusted_eod_prices").fetchone()[0]
    if not snapshot_date:
        raise ValueError("No adjusted price data available.")
    return snapshot_date


def load_snapshot(conn, snapshot_date):
    query = """
        SELECT
            a.symbol,
            s.company_name,
            s.isin,
            s.series,
            s.instrument_type,
            a.close,
            m.market_cap_cr,
            m.shares_outstanding,
            i.ma_20 AS dma_20,
            i.ma_50 AS dma_50,
            i.ma_100 AS dma_100,
            i.ma_200 AS dma_200
        FROM adjusted_eod_prices a
        LEFT JOIN symbols s
            ON s.symbol = a.symbol
        LEFT JOIN marketcap m
            ON m.symbol = a.symbol AND m.date = a.date
        LEFT JOIN indicators i
            ON i.symbol = a.symbol AND i.date = a.date
        WHERE a.date = ?
          AND a.close > 0
    """
    snapshot = pd.read_sql(query, conn, params=[snapshot_date])
    if snapshot.empty:
        return snapshot
    numeric_cols = ["close", "market_cap_cr", "shares_outstanding", "dma_20", "dma_50", "dma_100", "dma_200"]
    for col in numeric_cols:
        if col in snapshot.columns:
            snapshot[col] = pd.to_numeric(snapshot[col], errors="coerce")
    return snapshot


def apply_mcap_filter(snapshot_df, minimum_mcap_cr):
    if snapshot_df.empty:
        return snapshot_df
    filtered = snapshot_df[
        snapshot_df["market_cap_cr"].notna() &
        (snapshot_df["market_cap_cr"] >= minimum_mcap_cr)
    ].copy()
    return filtered


def load_price_history(conn, symbols, as_of_date=None, days=DAYS_TO_LOAD):
    if not symbols:
        return pd.DataFrame(columns=["symbol", "date", "close", "volume"])
    frames = []
    batch_size = 900
    for start in range(0, len(symbols), batch_size):
        batch = symbols[start:start + batch_size]
        placeholders = ",".join("?" for _ in batch)
        params = list(batch)
        date_clause = ""
        if as_of_date:
            date_clause = "AND date <= ?"
            params.append(as_of_date)
        query = f"""
            SELECT symbol, date, close, volume
            FROM (
                SELECT
                    symbol,
                    date,
                    close,
                    volume,
                    ROW_NUMBER() OVER (PARTITION BY symbol ORDER BY date DESC) AS rn
                FROM adjusted_eod_prices
                WHERE symbol IN ({placeholders})
                  AND close > 0
                  {date_clause}
            )
            WHERE rn <= {days}
        """
        frames.append(pd.read_sql(query, conn, params=params))
    history = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if history.empty:
        return history
    history["date"] = pd.to_datetime(history["date"])
    history["close"] = pd.to_numeric(history["close"], errors="coerce")
    history["volume"] = pd.to_numeric(history["volume"], errors="coerce")
    history.sort_values(["symbol", "date"], inplace=True)
    history.reset_index(drop=True, inplace=True)
    return history


def next_available_price(symbol_df, target_date):
    candidates = symbol_df[symbol_df["date"] >= target_date]
    if candidates.empty:
        return None, None
    first = candidates.iloc[0]
    return float(first["close"]), first["date"]


def compute_annual_roc_filter(prices_df, snapshot_date, minimum_roc):
    if prices_df.empty:
        return pd.DataFrame(columns=["symbol", "ROC_annual"])
    target_date = pd.to_datetime(snapshot_date) - pd.DateOffset(months=12)
    rows = []
    for symbol, group in prices_df.groupby("symbol"):
        group = group.sort_values("date")
        latest_close = float(group.iloc[-1]["close"])
        base_close, base_date = next_available_price(group, target_date)
        if base_close is None or base_close <= 0:
            continue
        if (base_date - target_date).days > 15:
            continue
        roc_value = (latest_close - base_close) / base_close * 100.0
        if roc_value >= minimum_roc:
            rows.append({"symbol": symbol, "ROC_annual": round2(roc_value)})
    return pd.DataFrame(rows)


def compute_turnover_filter(prices_df, symbols, minimum_turnover_cr):
    if prices_df.empty or not symbols:
        return pd.DataFrame(columns=["symbol", "median_turnover_cr"])
    subset = prices_df[prices_df["symbol"].isin(symbols)].copy()
    rows = []
    for symbol, group in subset.groupby("symbol"):
        recent = group.sort_values("date").tail(TRADING_DAYS_52W)
        if len(recent) < 100:
            continue
        turnover_cr = np.median(recent["close"].values * recent["volume"].fillna(0).values) / 1e7
        if turnover_cr >= minimum_turnover_cr:
            rows.append({"symbol": symbol, "median_turnover_cr": round2(turnover_cr)})
    return pd.DataFrame(rows)


def compute_sharpe_metrics(prices_df, symbols, snapshot_date, long_months=6, short_months=3):
    if prices_df.empty or not symbols:
        return pd.DataFrame(columns=["symbol"])
    long_target = pd.to_datetime(snapshot_date) - pd.DateOffset(months=long_months)
    short_target = pd.to_datetime(snapshot_date) - pd.DateOffset(months=short_months)
    rows = []
    subset = prices_df[prices_df["symbol"].isin(symbols)].copy()
    for symbol, group in subset.groupby("symbol"):
        group = group.sort_values("date")
        latest_close = float(group.iloc[-1]["close"])

        sharpe_long = None
        roc_long = None
        sharpe_long_raw = None
        base_close_long, base_date_long = next_available_price(group, long_target)
        if base_close_long and base_close_long > 0 and (base_date_long - long_target).days <= 15:
            long_window = group[group["date"] >= base_date_long]["close"].astype(float).to_numpy()
            if len(long_window) > 2:
                returns = np.diff(long_window) / long_window[:-1]
                std = float(np.std(returns, ddof=1))
                if std > 0:
                    sharpe_long_raw = float(np.mean(returns) / std)
                    sharpe_long = round2(sharpe_long_raw)
                roc_long = round2((latest_close - base_close_long) / base_close_long * 100.0)

        sharpe_short = None
        roc_short = None
        sharpe_short_raw = None
        base_close_short, base_date_short = next_available_price(group, short_target)
        if base_close_short and base_close_short > 0 and (base_date_short - short_target).days <= 10:
            short_window = group[group["date"] >= base_date_short]["close"].astype(float).to_numpy()
            if len(short_window) > 2:
                returns = np.diff(short_window) / short_window[:-1]
                std = float(np.std(returns, ddof=1))
                if std > 0:
                    sharpe_short_raw = float(np.mean(returns) / std)
                    sharpe_short = round2(sharpe_short_raw)
                roc_short = round2((latest_close - base_close_short) / base_close_short * 100.0)

        if sharpe_long_raw is None and sharpe_short_raw is None:
            continue

        week_52_high = None
        away_52wh = None
        if len(group) >= TRADING_DAYS_52W:
            week_52_high = round2(group["close"].tail(TRADING_DAYS_52W).max())
            if week_52_high and week_52_high > 0:
                away_52wh = round2((latest_close - week_52_high) / week_52_high * 100.0)

        rows.append({
            "symbol": symbol,
            "sharpe_long_raw": sharpe_long_raw,
            "sharpe_short_raw": sharpe_short_raw,
            "sharpe_6": sharpe_long,
            "sharpe_3": sharpe_short,
            "ROC_6": roc_long,
            "ROC_3": roc_short,
            "week_52_high": week_52_high,
            "away_52wh": away_52wh,
        })
    return pd.DataFrame(rows)


def compute_circuit_hits(prices_df, symbols, lookback_days=CIRCUIT_LOOKBACK):
    if prices_df.empty or not symbols:
        return pd.DataFrame(columns=["symbol", "total_circuit_hits_3m"])
    rows = []
    subset = prices_df[prices_df["symbol"].isin(symbols)].copy()
    for symbol, group in subset.groupby("symbol"):
        closes = group.sort_values("date")["close"].astype(float).to_numpy()
        window = closes[-(lookback_days + 1):]
        if len(window) < 2:
            continue
        previous = window[:-1]
        current = window[1:]
        valid = previous > 0
        returns = np.where(valid, (current - previous) / previous * 100.0, np.nan)
        hits = 0
        for change in returns:
            if np.isnan(change):
                continue
            for band in CIRCUIT_BANDS:
                if (band - CIRCUIT_TOLERANCE) <= change <= (band + CIRCUIT_TOLERANCE):
                    hits += 1
                    break
                if -(band + CIRCUIT_TOLERANCE) <= change <= -(band - CIRCUIT_TOLERANCE):
                    hits += 1
                    break
        rows.append({"symbol": symbol, "total_circuit_hits_3m": int(hits)})
    return pd.DataFrame(rows)


def rank_results(df):
    if df.empty:
        return df
    ranked = df.dropna(subset=["sharpe_long_raw", "sharpe_short_raw"]).copy()
    if ranked.empty:
        return ranked
    ranked["sharpe_6_rank"] = ranked["sharpe_long_raw"].rank(ascending=False, method="first").astype(int)
    ranked["sharpe_3_rank"] = ranked["sharpe_short_raw"].rank(ascending=False, method="first").astype(int)
    ranked["Avg_sharpe_6_3_Rank"] = ranked["sharpe_6_rank"] + ranked["sharpe_3_rank"]
    ranked.sort_values(["Avg_sharpe_6_3_Rank", "symbol"], inplace=True)
    ranked.reset_index(drop=True, inplace=True)
    return ranked


def apply_filtered_sheet_rules(df):
    if df.empty:
        return df
    filtered = df.copy()
    filtered = filtered[filtered["ROC_3"].notna() & (filtered["ROC_3"] > ROC_3M_FILTER)]
    filtered = filtered[filtered["week_52_high"].notna() & (filtered["close"] >= filtered["week_52_high"] * 0.75)]
    if "total_circuit_hits_3m" in filtered.columns:
        filtered = filtered[filtered["total_circuit_hits_3m"] <= CIRCUIT_MAX_HITS]
    return filtered.reset_index(drop=True)


def run_screener(mcap_filter=MCAP_FILTER_CR, roc_filter=ROC_ANNUAL_FILTER, turnover_filter=TURNOVER_FILTER_CR, as_of_date=None, long_months=6, short_months=3):
    if long_months <= 0 or short_months <= 0:
        raise ValueError("Sharpe month windows must be positive integers.")
    if long_months > 12 or short_months > 12:
        raise ValueError("Sharpe month windows cannot exceed 12.")

    with get_connection() as conn:
        setup_schema(conn)
        snapshot_date = resolve_snapshot_date(conn, as_of_date=as_of_date)
        snapshot = load_snapshot(conn, snapshot_date)
        screened = apply_mcap_filter(snapshot, mcap_filter)
        if screened.empty:
            return pd.DataFrame(), snapshot_date, pd.DataFrame()
        price_history = load_price_history(conn, screened["symbol"].tolist(), as_of_date=snapshot_date)

    if price_history.empty:
        return pd.DataFrame(), snapshot_date, price_history

    annual_roc = compute_annual_roc_filter(price_history, snapshot_date, roc_filter)
    if annual_roc.empty:
        return pd.DataFrame(), snapshot_date, price_history

    turnover = compute_turnover_filter(price_history, annual_roc["symbol"].tolist(), turnover_filter)
    if turnover.empty:
        return pd.DataFrame(), snapshot_date, price_history

    sharpe = compute_sharpe_metrics(
        price_history,
        turnover["symbol"].tolist(),
        snapshot_date,
        long_months=long_months,
        short_months=short_months,
    )
    if sharpe.empty:
        return pd.DataFrame(), snapshot_date, price_history

    result = (
        screened
        .merge(annual_roc, on="symbol", how="inner")
        .merge(turnover, on="symbol", how="inner")
        .merge(sharpe, on="symbol", how="inner")
    )
    result = rank_results(result)
    if result.empty:
        return result, snapshot_date, price_history

    result.attrs["long_months"] = int(long_months)
    result.attrs["short_months"] = int(short_months)
    output_cols = [
        "symbol", "company_name", "series", "instrument_type",
        "close", "dma_20", "dma_50", "dma_100", "dma_200",
        "away_52wh", "Avg_sharpe_6_3_Rank", "sharpe_6", "sharpe_3",
        "ROC_6", "ROC_3", "week_52_high", "market_cap_cr", "ROC_annual",
        "median_turnover_cr", "sharpe_6_rank", "sharpe_3_rank",
        "isin", "shares_outstanding",
    ]
    result = result[[col for col in output_cols if col in result.columns]]
    return result, snapshot_date, price_history


def print_results(df, top_n=TOP_N):
    if df.empty:
        print("No results to display.")
        return
    long_months = int(df.attrs.get("long_months", 6))
    short_months = int(df.attrs.get("short_months", 3))
    display = df.head(top_n).copy()
    display.index = range(1, len(display) + 1)
    cols = [
        "symbol", "company_name", "market_cap_cr", "ROC_annual", "median_turnover_cr",
        "sharpe_6", "sharpe_6_rank", "sharpe_3", "sharpe_3_rank", "Avg_sharpe_6_3_Rank",
    ]
    cols = [col for col in cols if col in display.columns]
    print("\n" + "=" * 90)
    print(f"Sharpe Screener | Sharpe {long_months}M/{short_months}M | Top {len(display)}")
    print("=" * 90)
    print(display[cols].to_string())
    print("=" * 90)


def export_to_excel(result, snapshot_date, price_history):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export the Sharpe screener workbook.") from exc

    if result.empty:
        return None

    all_rows = result.copy()
    circuit_df = compute_circuit_hits(price_history, all_rows["symbol"].tolist())
    all_rows = all_rows.merge(circuit_df, on="symbol", how="left")
    all_rows["total_circuit_hits_3m"] = all_rows["total_circuit_hits_3m"].fillna(0).astype(int)
    filtered_rows = apply_filtered_sheet_rules(all_rows)

    header_fill = PatternFill("solid", start_color="1F4E79")
    filter_fill = PatternFill("solid", start_color="1A5C38")
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    alt_fill_green = PatternFill("solid", start_color="D4EDDA")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    widths = {
        "symbol": 12, "company_name": 28, "series": 8, "instrument_type": 10,
        "close": 10, "dma_20": 10, "dma_50": 10, "dma_100": 10, "dma_200": 10,
        "away_52wh": 14, "Avg_sharpe_6_3_Rank": 16, "sharpe_6": 11, "sharpe_3": 11,
        "ROC_6": 10, "ROC_3": 10, "week_52_high": 13, "market_cap_cr": 14,
        "ROC_annual": 12, "median_turnover_cr": 16, "sharpe_6_rank": 13,
        "sharpe_3_rank": 13, "total_circuit_hits_3m": 16, "isin": 16, "shares_outstanding": 18,
    }
    friendly = {
        "symbol": "Symbol",
        "company_name": "Company Name",
        "series": "Series",
        "instrument_type": "Type",
        "close": "Close (Rs.)",
        "dma_20": "20 DMA",
        "dma_50": "50 DMA",
        "dma_100": "100 DMA",
        "dma_200": "200 DMA",
        "away_52wh": "Away from 52WH %",
        "Avg_sharpe_6_3_Rank": f"Avg Sharpe {int(result.attrs.get('long_months', 6))}M/{int(result.attrs.get('short_months', 3))}M Rank",
        "sharpe_6": f"Sharpe {int(result.attrs.get('long_months', 6))}M",
        "sharpe_3": f"Sharpe {int(result.attrs.get('short_months', 3))}M",
        "ROC_6": f"{int(result.attrs.get('long_months', 6))}M ROC %",
        "ROC_3": f"{int(result.attrs.get('short_months', 3))}M ROC %",
        "week_52_high": "52W High (Rs.)",
        "market_cap_cr": "MCAP (Cr)",
        "ROC_annual": "Annual ROC %",
        "median_turnover_cr": "Med. Turnover (Cr)",
        "sharpe_6_rank": f"Sharpe {int(result.attrs.get('long_months', 6))}M Rank",
        "sharpe_3_rank": f"Sharpe {int(result.attrs.get('short_months', 3))}M Rank",
        "total_circuit_hits_3m": "Circuit Hits (3M)",
        "isin": "ISIN",
        "shares_outstanding": "Shares Outstanding",
    }
    pct_cols = {"ROC_annual", "ROC_6", "ROC_3", "away_52wh"}
    num_cols = {"close", "dma_20", "dma_50", "dma_100", "dma_200", "week_52_high", "market_cap_cr", "median_turnover_cr", "sharpe_6", "sharpe_3", "shares_outstanding"}
    int_cols = {"sharpe_6_rank", "sharpe_3_rank", "Avg_sharpe_6_3_Rank", "total_circuit_hits_3m"}

    def write_sheet(ws, data, title, fill_main, fill_alt):
        cols = list(data.columns)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        title_cell.fill = fill_main
        title_cell.alignment = center
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=ci, value=friendly.get(col, col))
            cell.font = header_font
            cell.fill = fill_main
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 12)
        for ri, (_, row) in enumerate(data.iterrows(), start=3):
            row_fill = fill_alt if ri % 2 == 0 else PatternFill()
            for ci, col in enumerate(cols, start=1):
                value = row[col]
                cell = ws.cell(row=ri, column=ci)
                cell.font = data_font
                cell.border = border
                cell.fill = row_fill
                if col in pct_cols:
                    cell.value = float(value) / 100.0 if pd.notna(value) else None
                    cell.number_format = "0.00%"
                    cell.alignment = center
                elif col in num_cols:
                    cell.value = round2(value) if pd.notna(value) else None
                    cell.number_format = "#,##0.00"
                    cell.alignment = center
                elif col in int_cols:
                    cell.value = int(value) if pd.notna(value) else None
                    cell.number_format = "0"
                    cell.alignment = center
                else:
                    cell.value = str(value) if pd.notna(value) else ""
                    cell.alignment = left
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"

    workbook = Workbook()
    first = workbook.active
    first.title = "All Stocks"
    write_sheet(first, all_rows, f"NSE Sharpe Screener | {snapshot_date}", header_fill, alt_fill)
    second = workbook.create_sheet(title="Filtered")
    if filtered_rows.empty:
        second.cell(row=1, column=1, value="No stocks matched the filtered-sheet rules.")
    else:
        write_sheet(
            second,
            filtered_rows,
            f"Filtered | ROC 3M > {ROC_3M_FILTER}% | Close >= 75% of 52WH | Circuit Hits <= {CIRCUIT_MAX_HITS}",
            filter_fill,
            alt_fill_green,
        )
    out_path = Path(__file__).parent / f"{snapshot_date}.xlsx"
    workbook.save(out_path)
    return out_path


def main():
    options = parse_args(sys.argv[1:])
    if options["date"]:
        try:
            datetime.strptime(options["date"], "%Y-%m-%d")
        except ValueError:
            print(f"Invalid --date format: {options['date']}. Use YYYY-MM-DD.")
            return
    result, snapshot_date, price_history = run_screener(
        mcap_filter=options["mcap"],
        roc_filter=options["rf"],
        turnover_filter=options["turnover"],
        as_of_date=options["date"],
        long_months=options["long_months"],
        short_months=options["short_months"],
    )
    if result.empty:
        log.warning("No stocks matched the screener rules.")
        return
    print_results(result, top_n=options["top"])
    path = export_to_excel(result, snapshot_date, price_history)
    if path:
        log.info(f"Sharpe Screener workbook saved to {path}")


if __name__ == "__main__":
    main()
